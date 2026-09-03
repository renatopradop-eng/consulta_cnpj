"""Funções auxiliares: achatar registros JSON aninhados e exportar para XLSX."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Colunas que nunca devem aparecer na tabela de resultados nem no XLSX.
# As entradas com índice (ex: "contato_email[1].email") são uma rede de
# segurança: normalmente nem chegam a existir, porque
# _resumir_listas_de_objetos já reduz contato_email/contato_telefonico a
# uma única coluna antes do achatamento.
COLUNAS_EXCLUIDAS = {
    "CNPJ",
    "situacao_cadastral.situacao_atual",
    "situacao_cadastral.motivo",
    "situacao_cadastral.data",
    "atividade_principal.codigo",
    "contato_email[1].email",
    "contato_telefonico[0].ddd",
    "contato_telefonico[1].ddd",
}

# Únicas colunas exibidas na tabela de resultados e no XLSX, nessa ordem.
# Cada item é uma lista de nomes/prefixos candidatos — o primeiro que bater
# com algum segmento do caminho da coluna (ex: "endereco.municipio" bate com
# "municipio") é usado. Ajuste aqui se quiser exibir mais colunas ou se a
# Casa dos Dados usar nomes de campo diferentes.
COLUNAS_PRIORITARIAS = [
    ["capital_social"],
    ["municipio"],
    ["contato_email", "email"],
    ["contato_telefonico", "telefone", "ddd", "celular", "whatsapp"],
    ["cnae_principal", "atividade_principal"],
]

# Rótulo amigável exibido na tabela/XLSX para cada grupo de COLUNAS_PRIORITARIAS
# (mesma posição/ordem). Os dados continuam indexados pelo nome técnico da
# coluna — isso só troca o texto do cabeçalho.
ROTULOS_PRIORITARIOS = [
    "Capital Social",
    "Município",
    "E-mail",
    "Telefone",
    "CNAE Principal",
]

# Campos que vêm como lista de objetos e devem virar uma única string legível
# por empresa (evita colunas fragmentadas por índice, tipo campo[0].chave,
# campo[1].chave...). Cada valor é a ordem de preferência de sub-chave para
# resumir cada item da lista.
CAMPOS_LISTA_PARA_RESUMIR = {
    "quadro_societario": ["nome", "nome_socio", "razao_social", "nome_completo"],
    "socios": ["nome", "nome_socio", "razao_social", "nome_completo"],
    "contato_telefonico": ["completo", "numero", "telefone"],
    "contato_email": ["email"],
    "telefones": ["numero", "telefone", "numero_telefone"],
    "emails": ["email", "endereco_email"],
}


def _resumir_item_lista(item, chaves_preferidas):
    if isinstance(item, dict):
        for chave in chaves_preferidas:
            if item.get(chave):
                return str(item[chave])
        valores = [str(v) for v in item.values() if v]
        return " ".join(valores)
    return str(item) if item is not None else ""


def _resumir_listas_de_objetos(records):
    """Reduz campos que vêm como lista de objetos (sócios, telefones, e-mails)
    a uma única string legível por empresa. Se o campo já vier como valor
    simples ou objeto único, não faz nada (fica para o flatten_record normal)."""
    preparados = []
    for record in records:
        if not isinstance(record, dict):
            preparados.append(record)
            continue
        record = dict(record)
        for chave, prioridades in CAMPOS_LISTA_PARA_RESUMIR.items():
            if isinstance(record.get(chave), list):
                record[chave] = " | ".join(
                    filter(None, (_resumir_item_lista(item, prioridades) for item in record[chave]))
                )
        preparados.append(record)
    return preparados


def _formatar_moeda_brl(valor):
    """Formata um número como moeda brasileira (R$ 1.234,56)."""
    try:
        numero = float(valor)
    except (TypeError, ValueError):
        return valor
    texto = f"{numero:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {texto}"


def _segmentos(coluna):
    return coluna.lower().replace("[", ".").replace("]", "").split(".")


def _bate_com_prioridade(coluna, candidatos):
    segmentos = _segmentos(coluna)
    return any(seg.startswith(candidato) for seg in segmentos for candidato in candidatos)


def _selecionar_e_ordenar_colunas(colunas):
    """Mantém apenas as colunas de COLUNAS_PRIORITARIAS, na ordem definida ali
    (as demais, incluindo COLUNAS_EXCLUIDAS, não aparecem na tabela/XLSX)."""
    restantes = [c for c in colunas if c not in COLUNAS_EXCLUIDAS]

    usadas = set()
    ordenadas = []
    for candidatos in COLUNAS_PRIORITARIAS:
        for col in restantes:
            if col not in usadas and _bate_com_prioridade(col, candidatos):
                ordenadas.append(col)
                usadas.add(col)

    return ordenadas


def rotulo_coluna(coluna):
    """Rótulo amigável para exibir no cabeçalho da coluna (tabela/XLSX).
    Os dados continuam indexados pelo nome técnico — isso só troca o texto."""
    segmentos = _segmentos(coluna)
    for candidatos, rotulo in zip(COLUNAS_PRIORITARIAS, ROTULOS_PRIORITARIOS):
        if any(seg.startswith(c) for seg in segmentos for c in candidatos):
            return rotulo
    return coluna


def flatten_record(record, parent_key="", sep="."):
    """Achata um dict aninhado em um único nível, unindo listas simples com ' | '."""
    items = {}
    if isinstance(record, dict):
        for key, value in record.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else str(key)
            items.update(flatten_record(value, new_key, sep))
    elif isinstance(record, list):
        if all(not isinstance(v, (dict, list)) for v in record):
            items[parent_key] = " | ".join(str(v) for v in record if v is not None)
        else:
            for idx, value in enumerate(record):
                items.update(flatten_record(value, f"{parent_key}[{idx}]", sep))
    else:
        items[parent_key] = record
    return items


def flatten_records(records):
    """Achata uma lista de registros e retorna (colunas_ordenadas, linhas_flat).

    Aplica _resumir_listas_de_objetos (uma coluna legível em vez de várias
    fragmentadas), formata capital_social como moeda e aplica
    _selecionar_e_ordenar_colunas (mantém só COLUNAS_PRIORITARIAS, remove
    COLUNAS_EXCLUIDAS) antes de definir a lista final de colunas.
    """
    records = _resumir_listas_de_objetos(records)
    flat_rows = [flatten_record(r) for r in records]
    for row in flat_rows:
        if "capital_social" in row:
            row["capital_social"] = _formatar_moeda_brl(row["capital_social"])
    colunas = []
    vistas = set()
    for row in flat_rows:
        for key in row.keys():
            if key not in vistas:
                vistas.add(key)
                colunas.append(key)
    colunas = _selecionar_e_ordenar_colunas(colunas)
    return colunas, flat_rows


def build_xlsx(records, sheet_name="Empresas"):
    """Gera um arquivo XLSX em memória a partir de uma lista de registros (dicts)."""
    colunas, flat_rows = flatten_records(records)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Empresas"

    rotulos = [rotulo_coluna(col) for col in colunas]
    ws.append(rotulos)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in flat_rows:
        ws.append([row.get(col, "") for col in colunas])

    for idx, (col, rotulo) in enumerate(zip(colunas, rotulos), start=1):
        max_len = max([len(rotulo)] + [len(str(row.get(col, ""))) for row in flat_rows]) if flat_rows else len(rotulo)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 60)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
